from pathlib import Path
from categorizer import (
    extract_subcategories,
    classify,
    score,
    extract_text,
    Categorizer,
)
from categorizer.validator import confirm
from openpyxl import Workbook  # type: ignore
from fastapi.testclient import TestClient
from file_classifier import app
import file_classifier
import sys
import json
import os


def test_score():
    assert score("bug bug", ["bug"]) == 1
    assert score("no match", ["bug"]) == 0


def test_extract_subcategories():
    text = "Lo speaker non funziona. Il bug causa problemi allo speaker."
    subs = extract_subcategories(text, max_terms=3)
    assert "speaker" in subs
    assert "bug" in subs


def test_classify():
    text = "manuale per speaker rotto"
    cat, subs, conf, amb = classify(text, "doc.txt")
    assert cat == "product_guide" or cat == "tech_assistance"
    assert isinstance(subs, list)
    assert conf > 0
    assert isinstance(amb, bool)


def test_confirmation_auto(monkeypatch):
    inputs = iter(["n", "product_price", "speaker,bug"])
    monkeypatch.setattr("builtins.input", lambda *_: next(inputs))
    category, subcats, validated, source, conf = confirm(
        "tech_assistance",
        ["speaker"],
        "bug nel sistema",
        "bug.txt",
        mode="interactive",
        confidence=0.4,
    )
    assert validated
    assert source == "human_override"


def test_categorizer_run(tmp_path: Path):
    sample = tmp_path / "help.txt"
    sample.write_text("Lo speaker non funziona, ho un problema. C'e' un bug.")
    cat = Categorizer(mode="auto")
    results = cat.run(tmp_path)
    assert results[0]["category"] == "tech_assistance"
    assert "speaker" in results[0]["subcategories"]
    assert "entities" in results[0]["metadata"]


def test_forced_category(tmp_path: Path):
    sample = tmp_path / "doc.txt"
    sample.write_text("Preventivo per il nuovo software")
    cat = Categorizer(mode="silent", main_category="product_price")
    results = cat.run(tmp_path)
    assert results[0]["category"] == "product_price"


def test_api_classify(tmp_path: Path):
    sample = tmp_path / "doc.txt"
    sample.write_text("Manuale di installazione")
    client = TestClient(app)
    resp = client.post(
        "/classify",
        json={"input_path": str(tmp_path), "category": "product_guide"},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data[0]["category"] == "product_guide"


def test_extract_xlsx(tmp_path: Path):
    sample = tmp_path / "data.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "hello"
    ws["A2"] = "world"
    wb.save(sample)
    text = extract_text(sample)
    assert "hello" in text
    assert "world" in text


def test_main_product_price_creates_prices(tmp_path: Path, monkeypatch):
    sample = tmp_path / "doc.txt"
    sample.write_text("test")

    sample_data = [{"file": str(sample), "category": "product_price"}]
    monkeypatch.setattr(
        "file_classifier.Categorizer.run", lambda self, path: sample_data
    )

    monkeypatch.setattr(
        sys,
        "argv",
        [
            "file_classifier.py",
            str(tmp_path),
            "--category",
            "product_price",
        ],
    )
    cwd = os.getcwd()
    os.chdir(tmp_path)
    try:
        file_classifier.main()
    finally:
        os.chdir(cwd)

    data = json.loads((tmp_path / "prices.json").read_text())
    assert data == sample_data
