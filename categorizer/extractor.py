from __future__ import annotations

import docx  # type: ignore
from openpyxl import load_workbook  # type: ignore
from bs4 import BeautifulSoup
from pathlib import Path
from pdfminer.high_level import extract_text as pdf_extract

from utils.logger import get_json_logger

logger = get_json_logger("extract_log")

_MIN_SIZE = 100  # bytes


def extract_text(path: Path) -> str:
    """Extract textual content from supported files, logging issues."""
    try:
        if path.suffix.lower() == ".pdf":
            text = pdf_extract(str(path))
        elif path.suffix.lower() == ".docx":
            doc = docx.Document(str(path))
            text = "\n".join(p.text for p in doc.paragraphs)
        elif path.suffix.lower() == ".xlsx":
            wb = load_workbook(filename=path, read_only=True, data_only=True)
            parts: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            parts.append(str(cell))
            text = "\n".join(parts)
        elif path.suffix.lower() in {".html", ".htm"}:
            soup = BeautifulSoup(path.read_text(), "html.parser")
            text = soup.get_text(" ", strip=True)
        else:
            text = path.read_text()
    except Exception as exc:
        logger.error("parse error", extra={"file": str(path), "error": repr(exc)})
        return ""

    if len(text.encode("utf-8")) < _MIN_SIZE:
        logger.warning("extracted text too short", extra={"file": str(path)})
    return text
