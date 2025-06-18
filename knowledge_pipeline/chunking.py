import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Protocol

import pandas as pd

from openpyxl import load_workbook
from .config import PipelineConfig
from .logging_config import setup_logging
from .llm_utils import (
    validate_record_with_llm,
    primary_llm_client,
)

logger = setup_logging()


class RecursiveCharacterTextSplitter:
    """Semplice splitter di testo ricorsivo basato su caratteri."""

    def __init__(self, chunk_size: int = 1024, chunk_overlap: int = 128, separators: Optional[List[str]] = None) -> None:
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.separators = separators or ["\n\n", "\n", " ", ""]

    def split_text(self, text: str) -> List[str]:
        final_chunks: List[str] = []
        queue = [text]
        for separator in self.separators:
            new_queue: List[str] = []
            for current_text in queue:
                if len(current_text) <= self.chunk_size:
                    new_queue.append(current_text)
                    continue
                parts = current_text.split(separator)
                temp_chunk: List[str] = []
                for part in parts:
                    if len(separator.join(temp_chunk + [part])) <= self.chunk_size:
                        temp_chunk.append(part)
                    else:
                        if temp_chunk:
                            new_queue.append(separator.join(temp_chunk))
                        temp_chunk = [part]
                if temp_chunk:
                    new_queue.append(separator.join(temp_chunk))
            queue = new_queue
        for chunk_candidate in queue:
            if len(chunk_candidate) > self.chunk_size:
                final_chunks.extend(self._split_with_overlap(chunk_candidate))
            else:
                final_chunks.append(chunk_candidate)
        return final_chunks

    def _split_with_overlap(self, text: str) -> List[str]:
        if len(text) <= self.chunk_size:
            return [text]
        chunks = []
        start = 0
        while start < len(text):
            end = min(start + self.chunk_size, len(text))
            chunks.append(text[start:end])
            if end == len(text):
                break
            start += self.chunk_size - self.chunk_overlap
            start = min(start, len(text) - 1)
            if self.chunk_overlap > 0 and start == end:
                break
        return chunks


class ChunkingStrategy(Protocol):
    def chunk(self, content: Any, source_id: str, metadata: Dict) -> List[Dict[str, Any]]: ...


class AdvancedSemanticChunker(ChunkingStrategy):
    def __init__(self, chunk_size: int, chunk_overlap: int) -> None:
        self.text_splitter = RecursiveCharacterTextSplitter(chunk_size=chunk_size, chunk_overlap=chunk_overlap)

    def chunk(self, text: str, source_id: str, metadata: Dict) -> List[Dict[str, Any]]:
        chunks = self.text_splitter.split_text(text)
        chunk_list = []
        for i, chunk_text in enumerate(chunks):
            chunk_metadata = dict(metadata)
            chunk_list.append(
                {
                    "chunk_id": f"{source_id}#chunk{i:04d}",
                    "source_document_id": source_id,
                    "content": chunk_text,
                    "metadata": chunk_metadata,
                }
            )
        return chunk_list

class StructuredDataExtractor(ChunkingStrategy):
    def chunk(self, file_path: Path, source_id: str, metadata: Dict) -> List[Dict[str, Any]]:
        try:
            wb = load_workbook(file_path, data_only=True)
            all_records: List[Dict[str, Any]] = []

            for sheet_name in wb.sheetnames:
                if sheet_name.strip().upper() in {"INDEX", "COVER", "SOMMARIO", "SUMMARY"}:
                    continue  # ignora fogli non rilevanti

                ws = wb[sheet_name]

                for row in ws.iter_rows(values_only=True):
                    row_values = [v for v in row if v is not None]
                    if not row_values:
                        continue

                    serial = None
                    description = None
                    price = None

                    # Estrai serial
                    for val in row_values:
                        if isinstance(val, str) and val.upper().startswith("GSP-"):
                            serial = val.strip()
                            break
                    if not serial:
                        continue

                    # Estrai il numero massimo plausibile come prezzo
                    numeric_vals = [float(v) for v in row_values if isinstance(v, (int, float)) and 1 <= v <= 10000]
                    price = max(numeric_vals) if numeric_vals else None

                    # Estrai descrizione testuale coerente
                    for val in row_values:
                        if isinstance(val, str) and val != serial and len(val.strip().split()) >= 2:
                            description = val.strip()
                            break

                    record = {
                        "serial": serial,
                        "description": description if description else None,
                        "price": price if price else None,
                        "sheet_name": sheet_name,
                        PipelineConfig.PRODUCT_STATUS_FIELD_NAME: (
                            "discontinued" if "discontinued" in sheet_name.lower() else "active"
                        )
                    }
                    validated_record = validate_record_with_llm(record)
                    all_records.append(validated_record)

            # Conversione in chunk semantico RAG-ready
            def render_as_text(rec: Dict[str, Any]) -> str:
                return (
                    f"Il componente {rec['serial']} è identificato come \"{rec['description']}\" "
                    f"con un prezzo di {rec['price']} €. "
                    f"È incluso nel foglio \"{rec['sheet_name']}\" ed è attualmente "
                    f"{'disponibile' if rec[PipelineConfig.PRODUCT_STATUS_FIELD_NAME] == 'active' else 'fuori produzione'}."
                )

            chunk_list = []
            for i, rec in enumerate(all_records):
                chunk_metadata = dict(metadata)
                chunk_list.append({
                "chunk_id": f"{source_id}#row{i:04d}",
                "source_document_id": source_id,
                "content": render_as_text(rec), 
                "metadata": chunk_metadata,
                })

            return chunk_list

        except Exception:
            logger.error(f"Errore durante l'estrazione strutturata da '{file_path.name}'", exc_info=True)
            return []